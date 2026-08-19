import logging
import io
import config

logger = logging.getLogger('FII.export')

def generate_pdf_report(top_df, params, projections, fig_bytes=None) -> bytes:
    """
    Gera um relatório PDF com os FIIs recomendados.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        logger.error("fpdf2 não instalado.")
        return b""
        
    try:
        class PDF(FPDF):
            def header(self):
                self.set_font("helvetica", "B", 15)
                self.cell(0, 10, "Relatorio de Alocacao de FIIs", new_x="LMARGIN", new_y="NEXT", align="C")
                
            def footer(self):
                self.set_y(-15)
                self.set_font("helvetica", "I", 8)
                self.cell(0, 10, f"Pagina {self.page_no()}", align="C")

        pdf = PDF()
        pdf.add_page()
        
        # Parâmetros
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(0, 10, "Parametros da Analise", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("helvetica", "", 10)
        param_labels = {
            "capital": "Capital a Investir (R$)",
            "profile": "Perfil do Investidor",
            "objective": "Objetivo Principal",
            "horizon_months": "Horizonte (meses)",
            "n_fiis": "Quantidade de FIIs na Carteira",
            "criterio_alocacao": "Critério de Alocação"
        }
        
        for k, v in params.items():
            if k in param_labels:
                val = f"R$ {v:,.2f}" if k == "capital" else str(v).title() if isinstance(v, str) else str(v)
                pdf.cell(0, 6, f"{param_labels[k]}: {val}", new_x="LMARGIN", new_y="NEXT")
            
        pdf.ln(5)
        
        # Sumário Executivo
        if top_df is not None and not top_df.empty:
            pdf.set_font("helvetica", "B", 12)
            pdf.cell(0, 10, "Sumario Executivo (Top Picks)", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("helvetica", "", 10)
            for i, (ticker, row) in enumerate(top_df.head(5).iterrows()):
                seg = str(row.get('Segmento', ''))
                adeq = row.get('Adequação', row.get('Adequação ao Perfil', ''))
                val = row.get('Valor_Alocado', row.get('Valor a Investir', ''))
                adeq_str = f"{adeq:.1%}" if isinstance(adeq, (int, float)) else str(adeq)
                val_str = f"R$ {val:,.2f}" if isinstance(val, (int, float)) else str(val)
                pdf.cell(0, 6, f"{i+1}. {ticker} ({seg}) - Adequacao: {adeq_str} | Alocacao: {val_str}", new_x="LMARGIN", new_y="NEXT")
            
        pdf.ln(5)
        
        # Disclaimer
        pdf.set_font("helvetica", "I", 8)
        disclaimer = getattr(config, 'DISCLAIMER_TEXT', "Relatório educacional. Não é recomendação de investimento.")
        # Remove emojis for standard pdf font
        disclaimer_clean = disclaimer.replace("⚠️", "").replace("🔒", "").replace("**", "")
        pdf.multi_cell(0, 5, disclaimer_clean)
        
        return bytes(pdf.output())
    except Exception as e:
        logger.error(f"Erro ao gerar PDF: {e}")
        return b""

def generate_excel_report(display_df) -> bytes:
    """
    Gera um relatório Excel (workbook) com o ranking formatado.
    """
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        logger.error("openpyxl não instalado.")
        return b""
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranking FIIs"
    
    for r in dataframe_to_rows(display_df, index=True, header=True):
        ws.append(r)
        
    ws_params = wb.create_sheet(title="Parâmetros")
    ws_params.append(["Nota", "Parâmetros estão registrados no sistema e PDF."])
    
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def generate_csv_export(display_df) -> str:
    """
    Exporta o dataframe como string CSV.
    """
    return display_df.to_csv(index=True)
